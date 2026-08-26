import asyncio
import os
from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import CommandStart
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = os.getenv("ADMIN_ID")
PHONE_NUMBER = os.getenv("PHONE_NUMBER", "")
LATITUDE = float(os.getenv("LATITUDE", "0.0"))
LONGITUDE = float(os.getenv("LONGITUDE", "0.0"))
ADDRESS_TEXT = os.getenv("ADDRESS_TEXT", "")

router = Router()

class PropertyRequest(StatesGroup):
    name = State()
    category = State()
    preferences = State()
    phone = State()

CATEGORY_MAP = {
    "rent_living": "🔑 Rent Residential",
    "rent_comm": "🏬 Rent Commercial",
    "buy_living": "🏡 Buy Residential",
    "buy_comm": "🏢 Buy Commercial"
}

def init_excel_files():
    if not os.path.exists("real_estate_catalog.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["CategoryKey", "Property Title", "Price", "Description & Location"])
        
        # Demo data for each category
        ws.append(["rent_living", "🔑 Studio Apartment (Center)", "$800 / month", "Fully furnished studio, modern interior, near subway."])
        ws.append(["rent_living", "🔑 2-Bed Family House", "$1,500 / month", "2-story house with garage and private yard."])
        
        ws.append(["rent_comm", "🏬 Retail Shop (120 sq.m)", "$2,500 / month", "High foot traffic street, ground floor, showcase windows."])
        ws.append(["rent_comm", "🏬 Modern Office Floor", "$4,000 / month", "Business center, 300 sq.m, underground parking."])
        
        ws.append(["buy_living", "🏡 Luxury Villa in Suburbs", "$350,000", "4-bed villa, swimming pool, land plot 10 sq.m."])
        ws.append(["buy_living", "🏡 3-Room Apartment", "$120,000", "New building, 110 sq.m, panoramic city views."])
        
        ws.append(["buy_comm", "🏢 Warehouse & Logistics Center", "$600,000", "1500 sq.m storage space, heavy truck access, security."])
        ws.append(["buy_comm", "🏢 Commercial Building", "$1,200,000", "3-story freestanding building in commercial zone."])
        
        wb.save("real_estate_catalog.xlsx")

def save_to_excel(name, category, preferences, phone):
    file_name = "property_requests.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(["Client Name", "Category Interest", "Preferences & Budget", "Phone Number"])
    else:
        wb = load_workbook(file_name)
        ws = wb.active
    ws.append([name, category, preferences, phone])
    wb.save(file_name)

def get_main_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔑 Rent Residential", callback_data="cat_rent_living")],
        [InlineKeyboardButton(text="🏬 Rent Commercial", callback_data="cat_rent_comm")],
        [InlineKeyboardButton(text="🏡 Buy Residential", callback_data="cat_buy_living")],
        [InlineKeyboardButton(text="🏢 Buy Commercial", callback_data="cat_buy_comm")],
        [InlineKeyboardButton(text="🔍 Request Property Match", callback_data="menu_request")],
        [InlineKeyboardButton(text="📍 Contact & Location", callback_data="menu_location")]
    ])

def get_category_menu(category_key):
    init_excel_files()
    wb = load_workbook("real_estate_catalog.xlsx")
    ws = wb.active
    
    keyboard = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if row and len(row) >= 3 and row[0] == category_key:
            title = row[1]
            price = row[2]
            keyboard.append([InlineKeyboardButton(text=f"{title} — {price}", callback_data=f"prop_{idx}")])
            
    if not keyboard:
        keyboard.append([InlineKeyboardButton(text="No properties in this category yet", callback_data="none")])
        
    keyboard.append([InlineKeyboardButton(text="🔙 Back to Main Menu", callback_data="menu_main")])
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

@router.message(CommandStart())
async def start_cmd(message: Message):
    welcome_text = "Welcome to Premium Real Estate Agency! 🏡\nPlease select a property category below:"
    await message.answer(welcome_text, reply_markup=get_main_menu())

@router.callback_query(F.data == "menu_main")
async def back_to_main(call: CallbackQuery):
    await call.answer()
    try:
        await call.message.edit_text("Main Menu. Please select a property category:", reply_markup=get_main_menu())
    except Exception:
        pass

@router.callback_query(F.data.startswith("cat_"))
async def show_category_properties(call: CallbackQuery):
    await call.answer()
    category_key = call.data.replace("cat_", "")
    category_name = CATEGORY_MAP.get(category_key, "Property Catalog")
    
    try:
        await call.message.edit_text(f"📍 **{category_name}**\nAvailable properties:", reply_markup=get_category_menu(category_key), parse_mode="Markdown")
    except Exception:
        pass

@router.callback_query(F.data.startswith("prop_"))
async def property_details(call: CallbackQuery):
    idx_str = call.data.split("_")[1]
    if idx_str.isdigit():
        idx = int(idx_str)
        init_excel_files()
        wb = load_workbook("real_estate_catalog.xlsx")
        ws = wb.active
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and len(r) >= 4]
        if 0 <= idx < len(rows):
            row = rows[idx]
            title = row[1]
            price = row[2]
            desc = row[3]
            await call.answer(f"{title}\nPrice: {price}\n\n{desc}", show_alert=True)
            return
    await call.answer("Property details not found.", show_alert=True)

@router.callback_query(F.data == "menu_location")
async def show_location(call: CallbackQuery):
    await call.answer()
    clean_phone = PHONE_NUMBER.replace(" ", "").replace("-", "").replace("+", "")
    await call.message.answer(f"📍 Agency Office: {ADDRESS_TEXT}\n📞 Agent Phone: {PHONE_NUMBER}")
    
    if LATITUDE != 0.0 and LONGITUDE != 0.0:
        await call.message.answer_location(latitude=LATITUDE, longitude=LONGITUDE)
        
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📱 Message an Agent", url=f"https://t.me/+{clean_phone}")]
    ])
    await call.message.answer("Contact our lead agent directly:", reply_markup=kb)

# --- PROPERTY MATCH REQUEST BLOCK (FSM) ---

@router.callback_query(F.data == "menu_request")
async def request_start(call: CallbackQuery, state: FSMContext):
    await call.answer()
    await call.message.answer("Personalized Property Match 🔍\nPlease enter your Full Name:")
    await state.set_state(PropertyRequest.name)

@router.message(PropertyRequest.name)
async def request_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer("What exactly are you looking for? (e.g., Rent an apartment, Buy an office):")
    await state.set_state(PropertyRequest.category)

@router.message(PropertyRequest.category)
async def request_category(message: Message, state: FSMContext):
    await state.update_data(category=message.text)
    await message.answer("Please specify your budget, preferred location, and key requirements:")
    await state.set_state(PropertyRequest.preferences)

@router.message(PropertyRequest.preferences)
async def request_preferences(message: Message, state: FSMContext):
    await state.update_data(preferences=message.text)
    await message.answer("Please provide your contact phone number:")
    await state.set_state(PropertyRequest.phone)

@router.message(PropertyRequest.phone)
async def request_phone(message: Message, state: FSMContext):
    data = await state.get_data()
    name = data.get("name")
    category = data.get("category")
    preferences = data.get("preferences")
    phone = message.text

    save_to_excel(name, category, preferences, phone)

    if ADMIN_ID:
        admin_text = (f"🏡 New Property Request!\n\n"
                      f"👤 Client: {name}\n"
                      f"🎯 Category: {category}\n"
                      f"📝 Budget/Preferences: {preferences}\n"
                      f"📞 Phone: {phone}")
        try:
            await message.bot.send_message(chat_id=int(ADMIN_ID), text=admin_text)
        except Exception:
            pass

    await message.answer("Thank you! 🏡 Your request has been received. Our agent will select the best options and contact you shortly.", reply_markup=get_main_menu())
    await state.clear()

async def main():
    init_excel_files()
    bot = Bot(token=TOKEN)
    dp = Dispatcher()
    dp.include_router(router)
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())